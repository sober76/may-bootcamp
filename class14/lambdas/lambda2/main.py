import json
import boto3
from openpyxl import Workbook, load_workbook
from io import BytesIO

def lambda_handler(event, context):
    s3 = boto3.client('s3')
    
    # Read Excel file from S3
    bucket_name = event.get('bucket', 'your-bucket')
    file_key = event.get('file_key', 'input.xlsx')
    
    try:
        # Download Excel file from S3
        response = s3.get_object(Bucket=bucket_name, Key=file_key)
        excel_data = response['Body'].read()
        
        # Load workbook from memory
        workbook = load_workbook(BytesIO(excel_data))
        sheet = workbook.active
        
        # Read data from Excel
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
            if row[0]:  # Skip empty rows
                data.append({
                    'name': row[0],
                    'value': row[1] or 0,
                    'category': row[2] or 'N/A'
                })
        
        # Create new workbook with processed data
        new_wb = Workbook()
        new_sheet = new_wb.active
        new_sheet.title = "Processed Data"
        
        # Add headers
        new_sheet.append(['Name', 'Value', 'Category', 'Status'])
        
        # Add processed data
        for item in data:
            status = 'High' if item['value'] > 100 else 'Low'
            new_sheet.append([item['name'], item['value'], item['category'], status])
        
        # Save to memory
        output = BytesIO()
        new_wb.save(output)
        output.seek(0)
        
        # Upload processed file back to S3
        output_key = f"processed_{file_key}"
        s3.put_object(
            Bucket=bucket_name,
            Key=output_key,
            Body=output.getvalue(),
            ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        return {
            'statusCode': 200,
            'body': json.dumps({
                'message': 'Excel file processed successfully',
                'records_processed': len(data),
                'output_file': output_key,
                'summary': {
                    'total_value': sum(item['value'] for item in data),
                    'high_value_count': len([item for item in data if item['value'] > 100])
                }
            })
        }
        
    except Exception as e:
        return {
            'statusCode': 500,
            'body': json.dumps({
                'error': str(e)
            })
        }